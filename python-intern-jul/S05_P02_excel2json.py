
from openpyxl import load_workbook
import json
file_name = './data.xlsx'

workbook =  load_workbook(file_name)

sheet = workbook['Sheet1']

question = dict()
i = 1
for row in sheet.iter_rows():
    # question
    que = row[0].value
    ans = row[1].value
    optionA = row[2].value
    optionB = row[3].value
    optionC = row[4].value
    optionD = row[5].value

    question[str(i)]=(dict({
        "question":que,
        "ans":ans,
        "optionA":optionA,
        "optionB":optionB,
        "optionC":optionC,
        "optionD":optionD,
    }))
    i+=1

# print(question)

data = json.dumps(question,indent=6)

print(data)

file = open("./output.json","w")
file.write(data)
file.close()